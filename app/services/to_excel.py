from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from typing import List, Optional
from datetime import datetime
from babel.dates import format_date
from io import BytesIO

from app.models.admission import Admission
from app.schemas.excel import ApplicantFilterCourseSummary, ApplicantFilterExcel


status_color_map = {
    "02 - ยื่นใบสมัครแล้ว": "38761d", # เขียว
    "03 - รอพิจารณา": "ffc000",
    "04 - ผ่านการพิจารณา": "38761d",
    "05 - ไม่ผ่านการพิจารณา": "980000",

    "02 - รอตรวจสอบเอกสาร": "ffc000", # เหลือง
    "03 - เอกสารครบถ้วน": "38761d",
    "04 - เอกสารไม่ครบถ้วน": "980000", # แดง

    "02 - รอตรวจการชำระเงิน": "ffc000",
    "03 - ชำระเงินเรียบร้อย": "38761d"
}

thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
)


def create_excel(data: List[dict], file_name: str):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "หน้ารายชื่อผู้สมัคร"

    ws1.cell(row=1, column=1, value="รายชื่อผู้สมัครหลักสูตร ITCS/B - รอบ 1/68 - ICT Portfolio")
    ws1.cell(row=2, column=1, value="วันที่ส่งออก: 10 กุมภาพันธ์ 2025")
    ws1.cell(row=3, column=1, value="หลักสูตร: ITCS/B")
    ws1.cell(row=4, column=1, value="รอบสมัคร: 1/68 - ICT Portfolio")
    ws1.cell(row=5, column=1, value="สถานะการสมัคร: ทั้งหมด")
    ws1.cell(row=6, column=1, value="สถานะการชำระเงิน: ทั้งหมด")
    ws1.cell(row=7, column=1, value="จำนวนผู้สมัคร: 9 คน")
    
    headers = ["ลำดับที่", "ชื่อ", "นามสกุล", "วันเกิด"]
    
    for col, header in enumerate(headers, start=1):
        ws1.cell(row=9, column=col, value=header).font = Font(bold=True)

    for row, item in enumerate(data, start=10):
        ws1.cell(row=row, column=1, value=row-9)   
        ws1.cell(row=row, column=2, value=item.firstnameTH)
        ws1.cell(row=row, column=3, value=item.lastnameTH)
        ws1.cell(row=row, column=4, value=item.birthDate)

    ws2 = wb.create_sheet(title="หน้าจัดหลุ่มประเมิณเบื้องต้น")
    ws2.cell(row=1, column=1, value="อันนี้คือหน้าจัดหลุ่มนะจร๊ะ หรือก็คือหน้า 2 นั่นเอง")

    wb.save(file_name)
    return f"{file_name} saved successfully."


def excel_applicant_list(applicant: List[dict], admission: Optional[Admission], filter_data: ApplicantFilterExcel):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "sheet 1"

    today = format_date(datetime.now(), format='d MMMM y', locale='th')

    ws1.cell(row=1, column=1, value=f"ราบชื่อผู้สมัครหลักสูตร {admission.program} - {admission.roundName}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=2, column=1, value=f"วันที่ส่งออก: {today}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=3, column=1, value=f"หลักสูตร: {admission.program}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=4, column=1, value=f"รอบสมัคร: {admission.roundName}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=5, column=1, value=f"สถานะการสมัคร: {filter_data.admitStatus if filter_data.admitStatus else "ทั้งหมด"}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=6, column=1, value=f"สถานะการชำระเงิน: {filter_data.paymentStatus if filter_data.paymentStatus else "ทั้งหมด"}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=7, column=1, value=f"จำนวนผู้สมัครทั้งหมด: {len(applicant)}").font = Font(bold=True, name="Sarabun")

    
    headers = ["ลำดับที่", "หลักสูตร", "รอบการสมัคร", "เลขที่สมัคร", "ชื่อ - นามสกุล ผู้สมัคร", "โรงเรียน", "CGPA", "GPA Math", "GPA English", "GPA Sc and Techno",
               "สถานะการสมัคร", "สถานะเอกสาร", "สถานะการชำระเงิน", "อีเมล", "โทรศัพท์", "วันที่สมัคร"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=9, column=col, value=header)
        cell.font = Font(bold=True, name="Sarabun")
        cell.border = thin_border
    

    for row, item in enumerate(applicant, start=10):
        for row_idx, item in enumerate(applicant, start=10):
            row_data = [
                row_idx - 9,
                admission.program,
                admission.roundName,
                item["applicantId"],
                item["name"],
                item["school"],
                item["cgpa"],
                item["dst_m"],
                item["dst_e"],
                item["dst_s"],
                item["admissionStatus"],
                item["docStatus"],
                item["paymentStatus"],
                item["email"],
                item["phone"],
                item["applyDate"]
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)

                font_color = "000000"

                if col_idx in [11, 12, 13]:
                    if isinstance(value, str) and value in status_color_map:
                        font_color = status_color_map[value]

                cell.font = Font(name="Sarabun", color=font_color)
                cell.border = thin_border


    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def excel_screening_group(applicant: List[dict], admission: Optional[Admission], filter_data: ApplicantFilterExcel):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "sheet 1"

    today = format_date(datetime.now(), format='d MMMM y', locale='th')

    ws1.cell(row=1, column=1, value=f"ราบชื่อผู้สมัครสำหรับการคัดกรองเบื้องต้น รอบ {admission.program} - {admission.roundName}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=2, column=1, value=f"วันที่ส่งออก: {today}").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=3, column=1, value=f"จำนวนผู้สมัคร: {len(applicant)}").font = Font(bold=True, name="Sarabun")


    headers = ["ลำดับที่", "หลักสูตร", "รอบการสมัคร", "เลขที่สมัคร", "ชื่อ - นามสกุล ผู้สมัคร", "โรงเรียน",
               "สถานะการสมัคร", "สถานะเอกสาร", "สถานะการชำระเงิน", "อีเมล", "โทรศัพท์", "กรรมการพิจารณา"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws1.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True, name="Sarabun")
        cell.border = thin_border

    for row, item in enumerate(applicant, start=6):
        for row_idx, item in enumerate(applicant, start=6):
            row_data = [
                row_idx - 5,
                item["course"],
                item["round"],
                item["applicantNumber"],
                item["name"],
                item["school"],
                item["admissionStatus"],
                item["docStatus"],
                item["paymentStatus"],
                item["email"],
                item["phone"],
                item["ccName"]
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)

                font_color = "000000"

                if col_idx in [7, 8, 9]:
                    if isinstance(value, str) and value in status_color_map:
                        font_color = status_color_map[value]

                cell.font = Font(name="Sarabun", color=font_color)
                cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def excel_screening_summary(applicant: List[dict]):
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "สรุปผลการประเมิน"

    status_counter = Counter(item["admissionStatus"] for item in applicant)
    
    for row in range(1, 6):
        for col in range(1, 3):
            ws1.cell(row=row, column=col).border = thin_border

    ws1.cell(row=1, column=1, value="รายการ").font = Font(bold=True, name="Sarabun")
    ws1.cell(row=1, column=2, value="จำนวน (คน)").font = Font(bold=True, name="Sarabun")
    
    ws1.cell(row=2, column=1, value="จำนวนผู้สมัครที่อยู่ในรายการนี้").font = Font(name="Sarabun")
    ws1.cell(row=2, column=2, value=len(applicant)).font = Font(name="Sarabun")

    ws1.cell(row=3, column=1, value="ผ่านการพิจารณา").font = Font(name="Sarabun")
    ws1.cell(row=3, column=2, value=status_counter.get("04 - ผ่านการพิจารณา", 0)).font = Font(name="Sarabun")

    ws1.cell(row=4, column=1, value="ไม่ผ่านการพิจารณา").font = Font(name="Sarabun")
    ws1.cell(row=4, column=2, value=status_counter.get("05 - ไม่ผ่านการพิจารณา", 0)).font = Font(name="Sarabun")

    ws1.cell(row=5, column=1, value="รอพิจารณา").font = Font(name="Sarabun")
    ws1.cell(row=5, column=2, value=status_counter.get("03 - รอพิจารณา", 0)).font = Font(name="Sarabun")


    ws2 = wb.create_sheet(title="รายละเอียดผู้สมัคร")

    ws2.cell(row=1, column=1, value="รายงานผลการประเมินผู้สมัครเบื้องต้น").font = Font(bold=True, name="Sarabun")
    ws2.cell(row=2, column=1, value=f"จำนวนผู้สมัคร: {len(applicant)} คน").font = Font(bold=True, name="Sarabun")

    headers = ["ลำดับที่", "หลักสูตร", "รอบการสมัคร", "เลขที่สมัคร", "ชื่อ - นามสกุล ผู้สมัคร", "โรงเรียน", "CGPA", "GPA Math", "GPA English", "GPA Sc and Techno",
               "สถานะการสมัคร", "สถานะเอกสาร", "อีเมล", "โทรศัพท์", "กรรมการพิจารณา", "หมายเหตุ", "วันที่ประเมิน"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws2.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, name="Sarabun")
        cell.border = thin_border

    for row, item in enumerate(applicant, start=5):
        for row_idx, item in enumerate(applicant, start=5):
            row_data = [
                row_idx - 4,
                item["course"],
                item["round"],
                item["applicantNumber"],
                item["name"],
                item["school"],
                item["cgpa"],
                item["dst_m"],
                item["dst_e"],
                item["dst_s"],
                item["admissionStatus"],
                item["docStatus"],
                item["email"],
                item["phone"],
                item["ccName"],
                item["comment"],
                item["date"]
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)

                font_color = "000000"

                if col_idx in [10, 11]:
                    if isinstance(value, str) and value in status_color_map:
                        font_color = status_color_map[value]

                cell.font = Font(name="Sarabun", color=font_color)
                cell.border = thin_border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output












